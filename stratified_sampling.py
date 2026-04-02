import re, os, random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── CONFIG ──────────────────────────────────────────────────────────────────
TRANSCRIPT_DIR = "."        # change to your folder path if needed
OUTPUT_FILE    = "EEE_coding_sheet.xlsx"
TOTAL_PER_PARTICIPANT = 20
RANDOM_SEED    = 42
# ────────────────────────────────────────────────────────────────────────────

def parse_transcript(filepath):
    with open(filepath, 'r') as f:
        content = f.read()
    blocks = re.split(r'\n(\d{1,2}:\d{2}(?::\d{2})?)\n', content)
    utterances = []
    i = 1
    while i < len(blocks) - 1:
        timestamp = blocks[i].strip()
        text = blocks[i+1].strip()
        i += 2
        if not text:
            continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        ra_patterns = [
            r'^(mm-?hmm\.?|uh-?huh\.?|okay\.?|ok\.?|yes\.?|right\.?|alright\.?)$',
            r'^can you (tell|explain|describe)',
            r'^what (did|do|are|is|was)',
            r'^why (did|do|are|is|was)',
            r'^how (did|do|are|is|was)',
            r'^(and|so|but) (why|what|how)',
            r'^(no|not|nope|yeah|yep)[\.,]?$',
        ]
        participant_lines = []
        for line in lines:
            is_ra = False
            if len(line) < 80:
                for pat in ra_patterns:
                    if re.match(pat, line.lower()):
                        is_ra = True
                        break
            if not is_ra:
                participant_lines.append(line)
        if participant_lines:
            cleaned_text = ' '.join(participant_lines)
            if len(cleaned_text) > 20:
                utterances.append((timestamp, cleaned_text))
    return utterances

def stratified_sample(utterances, n):
    if len(utterances) <= n:
        return utterances
    indices = [int(round(i * (len(utterances) - 1) / (n - 1))) for i in range(n)]
    return [utterances[i] for i in sorted(set(indices))]

# Parse all gA and gB files
random.seed(RANDOM_SEED)
files = [f for f in os.listdir(TRANSCRIPT_DIR) 
         if f.endswith('.txt') and ('_gA_' in f or '_gB_' in f)]
files.sort()
print(f"Found {len(files)} transcript files (gA and gB only)")

participant_game_data = {}
for fname in files:
    match = re.match(r'(P\d+)_g([AB])_', fname)
    if not match:
        continue
    pid, game = match.group(1), match.group(2)
    fpath = os.path.join(TRANSCRIPT_DIR, fname)
    utterances = parse_transcript(fpath)
    if pid not in participant_game_data:
        participant_game_data[pid] = {}
    participant_game_data[pid][game] = utterances
    print(f"  {pid} Game {game}: {len(utterances)} utterances")

# Sample
rows = []
for pid in sorted(participant_game_data.keys()):
    games = participant_game_data[pid]
    total_utts = sum(len(v) for v in games.values())
    for game, utterances in sorted(games.items()):
        proportion = len(utterances) / total_utts
        n_sample = max(2, round(TOTAL_PER_PARTICIPANT * proportion))
        sampled = stratified_sample(utterances, n_sample)
        for ts, text in sampled:
            rows.append({
                'Participant ID': pid,
                'Game': game,
                'Timestamp': ts,
                'Utterance': text,
                'Manual Category': ''
            })

print(f"\nTotal utterances to code: {len(rows)}")
print(f"Participants: {len(participant_game_data)}")

# Build Excel
wb = Workbook()
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Coding Sheet ---
ws = wb.active
ws.title = "Coding Sheet"

headers = ['#', 'Participant ID', 'Game', 'Timestamp', 'Utterance', 'Manual Category']
col_widths = [5, 15, 8, 12, 80, 20]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 30

fill_light = PatternFill("solid", fgColor="F8F9FA")
fill_white = PatternFill("solid", fgColor="FFFFFF")
fill_cat   = PatternFill("solid", fgColor="EBF5FB")

for row_idx, row in enumerate(rows, 2):
    fill = fill_light if row_idx % 2 == 0 else fill_white
    values = [row_idx-1, row['Participant ID'], row['Game'], 
              row['Timestamp'], row['Utterance'], row['Manual Category']]
    aligns = ['center', 'center', 'center', 'center', 'left', 'center']
    for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal=align, vertical='top', wrap_text=True)
        cell.border = border
        cell.fill = fill_cat if col_idx == 6 else fill
    ws.row_dimensions[row_idx].height = 60

ws.freeze_panes = "A2"

dv = DataValidation(
    type="list",
    formula1='"Explore,Establish,Exploit,Unrelated,RA Speech"',
    allow_blank=True,
    showDropDown=False
)
dv.sqref = f"F2:F{len(rows)+1}"
ws.add_data_validation(dv)

# --- Category Definitions ---
ws2 = wb.create_sheet("Category Definitions")
defs = [
    ("Category", "Definition", "Key Verbal Markers"),
    ("Explore",
     "Participant is orienting to the environment, gathering information without a clear hypothesis. Actions are driven by uncertainty or curiosity rather than a specific prediction.",
     "what, why, how, maybe, might, could, not sure, don't know, trying to figure out, confused, hmm, let me see, looking, checking, just going to try, randomly, I think maybe, kind of, seems like"),
    ("Establish",
     "Participant is forming or testing a specific hypothesis. They have a candidate rule or causal belief and are actively checking whether it holds.",
     "I think, my hypothesis, if...then, probably, should be, let me test, testing, checking if, see if, trying to confirm, verify, gonna try, let me see if, assuming, suppose, expect, predict, I wonder if"),
    ("Exploit",
     "Participant has confirmed a rule and is applying it strategically. Actions are deliberate and directed toward the goal using known information.",
     "I know, definitely, clearly, for sure, figured it out, got it, understand, now I'll just, just need to, all I have to do, now I can, okay now, just, easy, almost done, finish"),
    ("Unrelated",
     "Utterance is not related to puzzle-solving reasoning (e.g., off-topic comment, meta-comment about the study).", "—"),
    ("RA Speech",
     "Utterance is from the Research Assistant, not the participant.", "—"),
]
cat_colors = {"Explore": "D6EAF8", "Establish": "D5F5E3", "Exploit": "FEF9E7", 
              "Unrelated": "F2F3F4", "RA Speech": "FADBD8"}
for col_idx, width in enumerate([18, 55, 65], 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
for row_idx, row_data in enumerate(defs, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border
        if row_idx == 1:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="2C3E50")
        else:
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=cat_colors.get(row_data[0], "FFFFFF"))
    ws2.row_dimensions[row_idx].height = 70 if row_idx > 1 else 25

# --- Summary ---
ws3 = wb.create_sheet("Summary")
ws3['A1'] = "Coding Progress Summary"
ws3['A1'].font = Font(name="Arial", bold=True, size=14)
ws3['A3'] = "Total Utterances"
ws3['B3'] = len(rows)
ws3['A4'] = "Coded"
ws3['B4'] = f"=COUNTA('Coding Sheet'!F2:F{len(rows)+1})"
ws3['A5'] = "Remaining"
ws3['B5'] = "=B3-B4"
ws3['A6'] = "% Complete"
ws3['B6'] = "=B4/B3"
ws3['B6'].number_format = '0.0%'
ws3['A8'] = "Category"
ws3['B8'] = "Count"
ws3['A8'].font = Font(bold=True)
ws3['B8'].font = Font(bold=True)
for i, cat in enumerate(["Explore", "Establish", "Exploit", "Unrelated", "RA Speech"], 9):
    ws3[f'A{i}'] = cat
    ws3[f'B{i}'] = f"=COUNTIF('Coding Sheet'!F:F,\"{cat}\")"
for col in ['A', 'B']:
    ws3.column_dimensions[col].width = 22

wb.save(OUTPUT_FILE)
print(f"\nSaved: {OUTPUT_FILE}")